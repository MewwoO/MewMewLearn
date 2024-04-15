"""
sss
"""

import os
import pandas as pd
from openpyxl import load_workbook


def process_excel_data(input_file, output_folder, first_row_content):
    """
    process excel data
    """
    # 创建保存文件的文件夹
    # pylint: disable=C0103
    # output_folder = "Day 01"
    os.makedirs(output_folder, exist_ok=True)

    # 读取表格数据
    df = pd.read_excel(input_file, header=None)

    # 删除第一行合并单元格的表格命名，保留列名
    df = df.drop(0)
    df = df.drop(1)

    # 重置索引
    df.reset_index(drop=True, inplace=True)

    # pylint: disable=C0103
    current_word_group = None
    word_group_rows = []
    current_phrase_group = None
    phrase_group_rows = []
    line = 0

    for _, row in df.iterrows():
        if not pd.isnull(row[1]) and not pd.isnull(row[2]):
            if len(word_group_rows) == 0:
                new_rows = [
                    [first_row_content],  # 第一行是合并了六个单元格的内容为test
                    [
                        "单词",
                        "英式音标",
                        "美式音标",
                        "释义",
                        "熟词僻义",
                        "助记",
                    ],  # 第二行是列名
                ]
                word_group_rows.extend(new_rows)
            word_group_rows.append(row)
            if len(phrase_group_rows) != 0:
                current_phrase_group = pd.DataFrame(phrase_group_rows)
                current_phrase_group.to_excel(
                    os.path.join(output_folder, f"{line}_phrase.xlsx"),
                    index=False,
                    header=False,
                )
                wb = load_workbook(output_folder + "/" + f"{line}_phrase.xlsx")
                ws = wb["Sheet1"]
                ws.merge_cells("A1:F1")
                wb.save(output_folder + "/" + f"{line}_phrase.xlsx")
                line += 1
                current_phrase_group = None
                phrase_group_rows = []
        elif pd.isnull(row[1]) and pd.isnull(row[2]):
            if len(phrase_group_rows) == 0:
                new_rows = [
                    [first_row_content],  # 第一行是合并了六个单元格的内容为test
                    [
                        "单词",
                        "英式音标",
                        "美式音标",
                        "释义",
                        "熟词僻义",
                        "助记",
                    ],  # 第二行是列名
                ]
                phrase_group_rows.extend(new_rows)
            phrase_group_rows.append(row)
            if len(word_group_rows) != 0:
                current_word_group = pd.DataFrame(word_group_rows)
                current_word_group.to_excel(
                    os.path.join(output_folder, f"{line}_word.xlsx"),
                    index=False,
                    header=False,
                )
                wb = load_workbook(output_folder + "/" + f"{line}_word.xlsx")
                ws = wb["Sheet1"]
                ws.merge_cells("A1:F1")
                wb.save(output_folder + "/" + f"{line}_word.xlsx")
                line += 1
                current_word_group = None
                word_group_rows = []

    if len(phrase_group_rows) != 0:
        current_phrase_group = pd.DataFrame(phrase_group_rows)
        current_phrase_group = pd.concat(
            [pd.DataFrame([df.iloc[0]]), current_word_group], ignore_index=True
        )
        current_phrase_group.to_excel(
            os.path.join(output_folder, f"{line}_phrase.xlsx"),
            index=False,
            header=False,
        )

    if len(word_group_rows) != 0:
        current_word_group = pd.DataFrame(word_group_rows)
        current_word_group = pd.concat(
            [pd.DataFrame([df.iloc[0]]), current_word_group], ignore_index=True
        )
        current_word_group.to_excel(
            os.path.join(output_folder, f"{line}_word.xlsx"), index=False, header=False
        )


process_excel_data(r"SparkCard\testPy\data.xlsx", "Day 01", "Mew_Word_Template")
